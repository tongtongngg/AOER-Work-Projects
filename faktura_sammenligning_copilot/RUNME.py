"""
Faktura-sammenligning RUNME — DTU AØR Copilot Agent

Formål
• Sammenligner to Excel-filer for en vilkårlig leverandør:
   - DOK 1: Excel fra leverandørens rykker-mail (forfaldne fakturaer)
   - DOK 2: Udtræk af leverandørfakturaer fra Fusion (Payables Report)
• Producerer ét output-Excel med 12 danske kolonner pr. faktura,
  plus en DATA ISSUES-fane med uoverensstemmelser og manglende felter.

Sådan bruges RUNME (Copilot)
1) Upload i samme besked:
   • Denne RUNME-fil
   • DOK 1 (rykker-mail Excel fra leverandøren)
   • DOK 2 (Fusion-udtræk Excel)
2) Skriv: Kør faktura-sammenligningen.
3) Copilot kalder run_pipeline(dok1_path, dok2_path, out_path) og leverer
   den færdige Excel.

Output-kolonner (i denne rækkefølge):
  Fakturanr., Leverandør navn, Status Modtaget, Faktura beløb,
  Faktura valuta, Fakturadato, Faktura modtaget dato i Fusion,
  Forfaldsdato leverandør, Fusion due date, Status Betalt,
  Status Godkendt, Afventer løft af hold.
"""

import os
import re
from datetime import datetime, date

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------
# Konfiguration
# ------------------------------------------------------------

DK_NUMBER = '#,##0.00'
DK_DATE = 'dd-mm-yyyy'

BANNER_FILL = "FFDDEBF7"
HEADER_FILL = "FFEAF2FF"
ISSUE_FILL = "FFFFEB9C"
MISSING_FILL = "FFFFC7CE"

OUTPUT_COLUMNS = [
    "Fakturanr.",
    "Leverandør navn",
    "Status Modtaget",
    "Faktura beløb",
    "Faktura valuta",
    "Fakturadato",
    "Faktura modtaget dato i Fusion",
    "Forfaldsdato leverandør",
    "Fusion due date",
    "Status Betalt",
    "Status Godkendt",
    "Afventer løft af hold",
]

DATE_OUTPUT_COLUMNS = {
    "Fakturadato",
    "Faktura modtaget dato i Fusion",
    "Forfaldsdato leverandør",
    "Fusion due date",
    "Status Betalt",  # kan indeholde dato
}

# Kandidat-headers for fleksibel kolonne-matching. Reelle eksporter har
# varierende stavemåde; vi normaliserer både headere og kandidater og
# laver eksakt match. Mere specifikke kandidater først.
COL_CANDIDATES = {
    "invoice_number":      ["invoice number", "document number", "fakturanr",
                            "fakturanummer", "invoice no", "faktura nr"],
    "supplier":            ["supplier", "supplier name", "leverandør", "leverandor",
                            "vendor", "vendor name", "leverandør navn"],
    "invoice_date":        ["invoice date", "fakturadato", "faktura dato"],
    "doc_date_dok1":       ["document date", "fakturadato", "faktura dato"],
    "invoice_amount":      ["invoice amount", "amount in local currency",
                            "faktura beløb", "beløb", "belob", "amount"],
    "invoice_currency":    ["invoice currency", "local currency",
                            "faktura valuta", "valuta", "currency"],
    "net_due_date":        ["net due date", "due date supplier",
                            "forfaldsdato leverandør", "forfaldsdato"],
    "payment_status":      ["payment status", "betalingsstatus"],
    "payment_date":        ["payment date", "betalingsdato", "paid date"],
    # Valgfrie Fusion-felter (ikke i sampleeksporten — log som DATA ISSUE hvis fraværende)
    "fusion_receipt_date": ["faktura modtaget dato i fusion", "invoice received date",
                            "received date", "receipt date", "fusion receipt date",
                            "modtaget dato"],
    "fusion_due_date":     ["fusion due date", "scheduled payment date",
                            "fusion forfaldsdato", "system due date"],
    "approval_status":     ["approval status", "status godkendt", "approved",
                            "godkendt", "approval"],
    "hold_status":         ["afventer løft af hold", "hold name", "hold status",
                            "on hold", "invoice hold", "hold"],
}

OPTIONAL_FUSION_FIELDS = [
    "fusion_receipt_date",
    "fusion_due_date",
    "approval_status",
    "hold_status",
]

DANISH_FIELD_LABEL = {
    "fusion_receipt_date": "Faktura modtaget dato i Fusion",
    "fusion_due_date": "Fusion due date",
    "approval_status": "Status Godkendt",
    "hold_status": "Afventer løft af hold",
}


# ------------------------------------------------------------
# Helpers
# ------------------------------------------------------------

def _normalize_header(s):
    if s is None:
        return ""
    s = str(s).lower().strip()
    s = re.sub(r"[._\-/]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


def find_column(df, candidates):
    """Return original column name matching any candidate (eksakt, normaliseret), else None."""
    norm_to_orig = {_normalize_header(c): c for c in df.columns}
    for cand in candidates:
        n = _normalize_header(cand)
        if n and n in norm_to_orig:
            return norm_to_orig[n]
    return None


def find_header_row(path, sheet_name, anchors, max_scan=15):
    """Scan first rows; return 0-indexed row where >=2 normalized anchors appear."""
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        norm_anchors = [_normalize_header(a) for a in anchors]
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
            cells = [_normalize_header(c) for c in row if c is not None]
            hits = sum(1 for a in norm_anchors if any(a == c for c in cells))
            if hits >= 2:
                return i
        return 0
    finally:
        wb.close()


def parse_number(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    s = s.replace(" ", "")
    if "," in s and (s.count(".") == 0 or s.rfind(",") > s.rfind(".")):
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        try:
            return float("".join(ch for ch in s if ch.isdigit() or ch in "+-."))
        except ValueError:
            return None


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, float) and np.isnan(val):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        ts = pd.to_datetime(val, dayfirst=True, errors="coerce")
        if pd.isna(ts):
            return None
        return ts.date()
    except Exception:
        return None


def normalize_invoice_number(x):
    """String join-key. Drops .0 from float-ified ints, preserves leading zeros, uppercases."""
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    s = str(x).strip()
    if not s:
        return ""
    if re.fullmatch(r"-?\d+\.0+", s):
        s = s.split(".", 1)[0]
    return s.upper()


def is_blank(v):
    if v is None:
        return True
    if isinstance(v, float) and np.isnan(v):
        return True
    if isinstance(v, str):
        return v.strip() == "" or v.strip().lower() in ("nan", "none", "nat")
    return False


def clean_str(v):
    return "" if is_blank(v) else str(v).strip()


def fmt_yes_no(b):
    return "ja" if b else "nej"


# ------------------------------------------------------------
# Loaders
# ------------------------------------------------------------

def _pick_data_sheet(path, min_rows=1):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for sn in wb.sheetnames:
            if wb[sn].max_row > min_rows:
                return sn
        return wb.sheetnames[0]
    finally:
        wb.close()


def load_dok1(path):
    """DOK 1 — leverandørens rykker-mail Excel."""
    sheet_name = _pick_data_sheet(path, min_rows=1)
    header_row = find_header_row(
        path, sheet_name,
        anchors=["document number", "invoice number", "document date",
                 "net due date", "amount in local currency"]
    )
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]

    col_inv  = find_column(df, COL_CANDIDATES["invoice_number"])
    col_date = find_column(df, COL_CANDIDATES["doc_date_dok1"])
    col_due  = find_column(df, COL_CANDIDATES["net_due_date"])
    col_amt  = find_column(df, COL_CANDIDATES["invoice_amount"])
    col_cur  = find_column(df, COL_CANDIDATES["invoice_currency"])

    if col_inv is None:
        raise ValueError(
            f"DOK 1: kunne ikke finde fakturanummer-kolonne. "
            f"Tilgængelige kolonner: {list(df.columns)}"
        )

    out = pd.DataFrame()
    out["__key"] = df[col_inv].apply(normalize_invoice_number)
    out["dok1_invoice_number"] = df[col_inv]
    out["dok1_document_date"] = df[col_date].apply(parse_date) if col_date else None
    out["dok1_net_due_date"] = df[col_due].apply(parse_date) if col_due else None
    out["dok1_amount"] = df[col_amt].apply(parse_number) if col_amt else None
    out["dok1_currency"] = df[col_cur].apply(clean_str) if col_cur else ""

    out = out[out["__key"] != ""].reset_index(drop=True)
    return out


def load_dok2(path):
    """DOK 2 — Fusion Payables Report. Returns (DataFrame, list of warnings)."""
    sheet_name = _pick_data_sheet(path, min_rows=3)
    header_row = find_header_row(
        path, sheet_name,
        anchors=["invoice number", "invoice date", "invoice amount",
                 "payment status", "supplier"]
    )
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row)
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]

    col_supplier = find_column(df, COL_CANDIDATES["supplier"])
    if col_supplier:
        # Forward-fill: DOK 2 har flettede celler i leverandørkolonnen
        df[col_supplier] = df[col_supplier].ffill()

    found = {}
    warnings = []
    for key, cands in COL_CANDIDATES.items():
        if key in ("doc_date_dok1", "net_due_date"):
            continue  # DOK 1-only felter
        col = find_column(df, cands)
        found[key] = col
        if col is None and key in OPTIONAL_FUSION_FIELDS:
            warnings.append({
                "Felt": DANISH_FIELD_LABEL.get(key, key),
                "Søgte navne": ", ".join(cands),
                "Note": ("Kolonnen blev ikke fundet i Fusion-eksporten. "
                         "Cellen efterlades tom — udvid Fusion-udtrækket "
                         "for at populere dette felt."),
            })

    if found["invoice_number"] is None:
        raise ValueError(
            f"DOK 2: kunne ikke finde fakturanummer-kolonne. "
            f"Tilgængelige kolonner: {list(df.columns)}"
        )

    out = pd.DataFrame()
    out["__key"] = df[found["invoice_number"]].apply(normalize_invoice_number)
    out["dok2_invoice_number"] = df[found["invoice_number"]]
    out["dok2_supplier"] = df[found["supplier"]].apply(clean_str) if found["supplier"] else ""
    out["dok2_invoice_date"] = df[found["invoice_date"]].apply(parse_date) if found["invoice_date"] else None
    out["dok2_amount"] = df[found["invoice_amount"]].apply(parse_number) if found["invoice_amount"] else None
    out["dok2_currency"] = df[found["invoice_currency"]].apply(clean_str) if found["invoice_currency"] else ""
    out["dok2_payment_status"] = df[found["payment_status"]].apply(clean_str) if found["payment_status"] else ""
    out["dok2_payment_date"] = df[found["payment_date"]].apply(parse_date) if found["payment_date"] else None
    out["dok2_fusion_receipt_date"] = df[found["fusion_receipt_date"]].apply(parse_date) if found["fusion_receipt_date"] else None
    out["dok2_fusion_due_date"] = df[found["fusion_due_date"]].apply(parse_date) if found["fusion_due_date"] else None
    out["dok2_approval_status"] = df[found["approval_status"]].apply(clean_str) if found["approval_status"] else None
    out["dok2_hold_status"] = df[found["hold_status"]].apply(clean_str) if found["hold_status"] else None

    # Markér om de valgfrie kolonner overhovedet findes (None = kolonnen mangler)
    out.attrs["has_approval_col"] = found["approval_status"] is not None
    out.attrs["has_hold_col"] = found["hold_status"] is not None

    out = out[out["__key"] != ""].reset_index(drop=True)
    return out, warnings


# ------------------------------------------------------------
# Output-bygning
# ------------------------------------------------------------

def _decide_status_betalt(payment_status, payment_date):
    """'Negotiable' eller tom => 'afventer betaling'. Ellers Payment Date hvis sat, ellers 'afventer betaling'."""
    ps = clean_str(payment_status).lower()
    if not ps or ps == "negotiable":
        return "afventer betaling"
    if payment_date is not None:
        return payment_date
    return "afventer betaling"


def _decide_godkendt(value, has_column):
    if not has_column:
        return ""  # kolonnen mangler — DATA ISSUE allerede logget
    s = clean_str(value).lower()
    if not s:
        return ""
    return fmt_yes_no(s in ("approved", "ja", "yes", "y", "true", "1", "godkendt"))


def _decide_afventer_hold(value, has_column):
    if not has_column:
        return ""
    s = clean_str(value).lower()
    if not s:
        return "nej"  # kolonnen findes, værdien er tom => intet hold
    if s in ("no", "nej", "false", "0", "none", "release", "released", "no hold"):
        return "nej"
    return "ja"  # ethvert hold-navn => afventer løft


def build_output(dok1_df, dok2_df, fallback_supplier_name):
    """Returns (main_df, dok1_only_rows, mismatches)."""
    has_approval_col = dok2_df.attrs.get("has_approval_col", False)
    has_hold_col = dok2_df.attrs.get("has_hold_col", False)

    merged = dok2_df.merge(dok1_df, on="__key", how="left", suffixes=("", "_d1"))

    rows = []
    mismatches = []
    for _, r in merged.iterrows():
        try:
            status_betalt = _decide_status_betalt(
                r.get("dok2_payment_status"), r.get("dok2_payment_date")
            )

            status_godkendt = _decide_godkendt(r.get("dok2_approval_status"), has_approval_col)
            afventer_hold = _decide_afventer_hold(r.get("dok2_hold_status"), has_hold_col)

            fakturadato = r.get("dok2_invoice_date") or r.get("dok1_document_date")

            leverandor = r.get("dok2_supplier")
            if is_blank(leverandor):
                leverandor = fallback_supplier_name

            valuta = clean_str(r.get("dok2_currency")) or clean_str(r.get("dok1_currency"))

            a1 = r.get("dok1_amount")
            a2 = r.get("dok2_amount")
            if a1 is not None and a2 is not None and abs(a1 - a2) > 0.01:
                mismatches.append({
                    "Fakturanr.": r["__key"],
                    "Beløb DOK 1": a1,
                    "Beløb DOK 2": a2,
                    "Difference (DOK2 - DOK1)": round(a2 - a1, 2),
                    "Note": "Beløb i rykker-mail (DOK 1) afviger fra Fusion (DOK 2) med >0,01.",
                })

            rows.append({
                "Fakturanr.": r["__key"],
                "Leverandør navn": leverandor,
                "Status Modtaget": "ja",
                "Faktura beløb": a2,
                "Faktura valuta": valuta,
                "Fakturadato": fakturadato,
                "Faktura modtaget dato i Fusion": r.get("dok2_fusion_receipt_date"),
                "Forfaldsdato leverandør": r.get("dok1_net_due_date"),
                "Fusion due date": r.get("dok2_fusion_due_date"),
                "Status Betalt": status_betalt,
                "Status Godkendt": status_godkendt,
                "Afventer løft af hold": afventer_hold,
            })
        except Exception as e:
            rows.append({
                "Fakturanr.": r.get("__key", ""),
                "Leverandør navn": f"FEJL: {e}",
                "Status Modtaget": "ja",
                "Faktura beløb": None,
                "Faktura valuta": "",
                "Fakturadato": None,
                "Faktura modtaget dato i Fusion": None,
                "Forfaldsdato leverandør": None,
                "Fusion due date": None,
                "Status Betalt": "",
                "Status Godkendt": "",
                "Afventer løft af hold": "",
            })

    main_df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)

    dok2_keys = set(dok2_df["__key"])
    dok1_only = dok1_df[~dok1_df["__key"].isin(dok2_keys)]
    dok1_only_rows = []
    for _, r in dok1_only.iterrows():
        dok1_only_rows.append({
            "Fakturanr.": r["__key"],
            "Fakturadato (DOK 1)": r.get("dok1_document_date"),
            "Forfaldsdato leverandør": r.get("dok1_net_due_date"),
            "Faktura beløb (DOK 1)": r.get("dok1_amount"),
            "Faktura valuta": r.get("dok1_currency"),
            "Note": "Findes i leverandørens rykker-mail, men ikke i Fusion-udtrækket. Status Modtaget = nej.",
        })

    return main_df, dok1_only_rows, mismatches


# ------------------------------------------------------------
# Excel-skriver
# ------------------------------------------------------------

def _bold_header_row(ws, row, headers):
    for j, c in enumerate(headers, 1):
        cell = ws.cell(row, j)
        cell.value = c
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type='solid')


def _autosize(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for c in col:
            if c.value is None:
                continue
            v = c.value
            if isinstance(v, (datetime, date)):
                s = "00-00-0000"
            else:
                s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def write_excel(out_path, main_df, warnings, dok1_only_rows, mismatches, banner):
    wb = Workbook()

    # ----- Hovedark -----
    ws = wb.active
    ws.title = "Faktura sammenligning"
    ws.append([banner])
    ws.cell(1, 1).fill = PatternFill(start_color=BANNER_FILL, end_color=BANNER_FILL, fill_type='solid')
    ws.cell(1, 1).font = Font(bold=True)
    _bold_header_row(ws, 2, OUTPUT_COLUMNS)

    for _, r in main_df.iterrows():
        ws.append([r.get(c) for c in OUTPUT_COLUMNS])

    hdr = {ws.cell(2, i).value: i for i in range(1, ws.max_column + 1)}
    for rr in range(3, ws.max_row + 1):
        if "Fakturanr." in hdr:
            ws.cell(rr, hdr["Fakturanr."]).number_format = '@'
        if "Faktura beløb" in hdr:
            cell = ws.cell(rr, hdr["Faktura beløb"])
            if isinstance(cell.value, (int, float)):
                cell.number_format = DK_NUMBER
        for cname in DATE_OUTPUT_COLUMNS:
            if cname in hdr:
                cell = ws.cell(rr, hdr[cname])
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = DK_DATE

    _autosize(ws)
    ws.freeze_panes = "A3"

    # ----- DATA ISSUES -----
    w2 = wb.create_sheet("DATA ISSUES")
    w2.append([banner])
    w2.cell(1, 1).fill = PatternFill(start_color=BANNER_FILL, end_color=BANNER_FILL, fill_type='solid')
    w2.cell(1, 1).font = Font(bold=True)

    cur = 2

    # Sektion 1 — manglende valgfrie kolonner
    w2.cell(cur, 1).value = "1) Manglende valgfrie kolonner i Fusion-eksport"
    w2.cell(cur, 1).font = Font(bold=True)
    cur += 1
    if warnings:
        cols = ["Felt", "Søgte navne", "Note"]
        _bold_header_row(w2, cur, cols)
        cur += 1
        for w in warnings:
            for j, c in enumerate(cols, 1):
                cell = w2.cell(cur, j)
                cell.value = w.get(c, "")
                cell.fill = PatternFill(start_color=MISSING_FILL, end_color=MISSING_FILL, fill_type='solid')
            cur += 1
    else:
        w2.cell(cur, 1).value = "Alle valgfrie kolonner blev fundet."
        cur += 1
    cur += 1

    # Sektion 2 — DOK1-only fakturaer
    w2.cell(cur, 1).value = "2) Fakturaer i leverandørens rykker-mail som mangler i Fusion (Status Modtaget = nej)"
    w2.cell(cur, 1).font = Font(bold=True)
    cur += 1
    if dok1_only_rows:
        cols = ["Fakturanr.", "Fakturadato (DOK 1)", "Forfaldsdato leverandør",
                "Faktura beløb (DOK 1)", "Faktura valuta", "Note"]
        _bold_header_row(w2, cur, cols)
        cur += 1
        for r in dok1_only_rows:
            for j, c in enumerate(cols, 1):
                cell = w2.cell(cur, j)
                cell.value = r.get(c)
                cell.fill = PatternFill(start_color=ISSUE_FILL, end_color=ISSUE_FILL, fill_type='solid')
                if c == "Fakturanr.":
                    cell.number_format = '@'
                elif c == "Faktura beløb (DOK 1)" and isinstance(cell.value, (int, float)):
                    cell.number_format = DK_NUMBER
                elif c in ("Fakturadato (DOK 1)", "Forfaldsdato leverandør") \
                        and isinstance(cell.value, (datetime, date)):
                    cell.number_format = DK_DATE
            cur += 1
    else:
        w2.cell(cur, 1).value = "Ingen — alle leverandørens rykker findes i Fusion."
        cur += 1
    cur += 1

    # Sektion 3 — beløb-mismatches
    w2.cell(cur, 1).value = "3) Beløb-uoverensstemmelser mellem DOK 1 og DOK 2"
    w2.cell(cur, 1).font = Font(bold=True)
    cur += 1
    if mismatches:
        cols = ["Fakturanr.", "Beløb DOK 1", "Beløb DOK 2", "Difference (DOK2 - DOK1)", "Note"]
        _bold_header_row(w2, cur, cols)
        cur += 1
        for m in mismatches:
            for j, c in enumerate(cols, 1):
                cell = w2.cell(cur, j)
                cell.value = m.get(c)
                cell.fill = PatternFill(start_color=ISSUE_FILL, end_color=ISSUE_FILL, fill_type='solid')
                if c == "Fakturanr.":
                    cell.number_format = '@'
                elif c in ("Beløb DOK 1", "Beløb DOK 2", "Difference (DOK2 - DOK1)") \
                        and isinstance(cell.value, (int, float)):
                    cell.number_format = DK_NUMBER
            cur += 1
    else:
        w2.cell(cur, 1).value = "Ingen beløb-uoverensstemmelser."
        cur += 1

    _autosize(w2)
    w2.freeze_panes = "A2"

    wb.save(out_path)


# ------------------------------------------------------------
# Pipeline
# ------------------------------------------------------------

def run_pipeline(dok1_path, dok2_path, out_path):
    """
    Hovedindgang. Copilot kalder denne funktion med:
      • dok1_path: Excel fra leverandørens rykker-mail
      • dok2_path: Fusion Payables-udtræk
      • out_path : ønsket sti til output-Excel
    Returnerer out_path.
    """
    dok1 = load_dok1(dok1_path)
    dok2, warnings = load_dok2(dok2_path)

    fallback_supplier = os.path.splitext(os.path.basename(dok1_path))[0]

    main_df, dok1_only_rows, mismatches = build_output(dok1, dok2, fallback_supplier)

    banner = (
        f"Faktura-sammenligning — DOK 1: {os.path.basename(dok1_path)} "
        f"— DOK 2: {os.path.basename(dok2_path)} "
        f"— genereret: {datetime.now():%Y-%m-%d %H:%M}"
    )

    write_excel(out_path, main_df, warnings, dok1_only_rows, mismatches, banner)
    return out_path

