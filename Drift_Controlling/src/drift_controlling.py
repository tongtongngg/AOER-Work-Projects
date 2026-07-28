#!/usr/bin/env python3

import argparse
import csv
import io
import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# True : en outputlinje pr. projekt + UK + udgiftskategori
# False: en outputlinje pr. projekt + udgiftskategori (UK'er lagt sammen)
GROUP_BY_UK = True

# True : hele linjen (kolonne A-O) faar roed skrift naar kolonne H er negativ
# False: kun selve H-cellen
RED_WHOLE_ROW = True

# Udgiftskategorier der regnes som drift. Alt andet frasorteres.
DRIFT_KATEGORIER = [
    "Intern produktion mv., køb",
    "Lejeomkostninger o.l.",
    "Rejser og repræsentation",
    "Øvrige omkostninger",
]

OUTPUT_HEADERS = [
    "Sektionsnummer",        # A
    "Projektnummer",         # B
    "Projektnavn",           # C
    "UK",                    # D
    "Udgiftskategori",       # E
    "Løbende budget YTD",    # F
    "Actuals YTD",           # G
    "Difference YTD",        # H
    "Difference YTD %",      # I
    "Benchmark",             # J
    "Beregnet justering",    # K
    "Kommentar YTD",         # L
    "Projektcontroller",     # M
    "Startdato",             # N
    "Slutdato",              # O
]

MAANEDSNAVNE = {
    1: "januar", 2: "februar", 3: "marts", 4: "april", 5: "maj", 6: "juni",
    7: "juli", 8: "august", 9: "september", 10: "oktober", 11: "november",
    12: "december",
}

FONT_NAME = "Arial"
RED = "FFC00000"
HEADER_FILL = PatternFill("solid", fgColor="FF1F3864")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFFFF00")


def read_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        return pd.read_excel(path, header=None, dtype=str)

    text = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if text is None:
        raise RuntimeError(f"Kunne ikke afkode {path.name}.")

    sample = "\n".join(text.splitlines()[:20])
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=";,\t").delimiter
    except csv.Error:
        delimiter = ";"

    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    rows = [r for r in rows if any(c.strip() for c in r)]
    if not rows:
        raise RuntimeError(f"{path.name} indeholder ingen raekker.")
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    return pd.DataFrame(rows, dtype=str)


def parse_dk_number(value) -> float:
    """'2.782.094,86' -> 2782094.86. Tomme felter -> 0.0"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in ("-", "nan"):
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace("\xa0", "").replace(" ", "")
    # Dansk format: '.' = tusindtal, ',' = decimal
    text = text.replace(".", "").replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -number if negative else number


def parse_us_date(value):
    """Projektstamdata leverer datoer som MM-DD-YYYY (fx 10-30-2023)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (pd.Timestamp, date)):
        return pd.Timestamp(value).date()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%m-%d-%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return pd.to_datetime(text, format=fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def extract_sektionsnummer(value) -> str:
    """'Afdeling for ... , (7810) Administration' -> '7810'"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value)
    match = re.search(r"\((\d{4})\)", text)
    if match:
        return match.group(1)
    match = re.search(r"(?<!\d)(\d{4})(?!\d)", text)
    return match.group(1) if match else ""


def clean(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


NUMERIC_RE = re.compile(r"^-?\(?[\d.\s\u00a0]*\d(?:,\d+)?\)?-?$")

# felt -> (forventet position jf. spec, moenstre der matcher kolonnenavnet)
LB_COLUMNS = {
    "projektnr":   (0, [r"project\s*no", r"projektnr"]),
    "projektnavn": (1, [r"projektnavn", r"project\s*name"]),
    "uk":          (2, [r"^uk$", r"underkonto"]),
    "kategori":    (6, [r"udgiftskategori", r"expenditure\s*category"]),
    "lb":          (7, [r"forecast\s*raw", r"l[øo]bende\s*budget"]),
    "actual":      (8, [r"actual\s*raw"]),
    "difference":  (9, [r"difference\s*raw"]),
}

SD_COLUMNS = {
    "projektnr":         (0, [r"projektnr"]),
    "projektnavn_sd":    (1, [r"projektnavn"]),
    "organisation":      (3, [r"project\s*organi", r"projektorgani"]),
    "projektstatus":     (4, [r"projektstatus"]),
    "startdato":         (5, [r"projektstartdato", r"startdato"]),
    "slutdato":          (6, [r"projektslutdato", r"slutdato"]),
    "projektcontroller": (8, [r"project\s*controller", r"projektcontroller"]),
}


def norm(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def looks_numeric(value) -> bool:
    text = norm(value)
    return bool(text) and bool(NUMERIC_RE.match(text)) and any(c.isdigit() for c in text)


def find_header_row(raw: pd.DataFrame, spec: dict, filename: str) -> int:
    """Headerraekken = den raekke hvor flest kolonnenavne genkendes."""
    best_row, best_hits = None, 0
    for i in range(min(25, len(raw))):
        cells = [norm(v) for v in raw.iloc[i]]
        hits = sum(
            any(re.search(p, c) for c in cells for p in patterns)
            for _, patterns in spec.values()
        )
        if hits > best_hits:
            best_row, best_hits = i, hits
    if best_row is None or best_hits < 3:
        raise RuntimeError(
            f"Kunne ikke finde headerraekken i {filename}. "
            f"Bedste kandidat genkendte {best_hits} af {len(spec)} kolonnenavne. "
            "Er filen semikolonsepareret og eksporteret fra den rigtige rapport?"
        )
    return best_row


def find_data_start(raw: pd.DataFrame, header_row: int, filename: str) -> int:
    for i in range(header_row + 1, len(raw)):
        if any(looks_numeric(v) for v in raw.iloc[i]):
            return i
    raise RuntimeError(
        f"Fandt ingen datalinjer i {filename} efter headerraekke {header_row + 1}."
    )


def map_columns(raw, header_row, data_start, spec, filename, warnings):
    """Byg sammensatte kolonnenavne og map felter til kolonneindeks."""
    composite = []
    for col in range(raw.shape[1]):
        parts = [clean(raw.iat[r, col]) for r in range(header_row, data_start)]
        composite.append(norm(" ".join(p for p in parts if p)))

    mapping = {}
    for field, (default_idx, patterns) in spec.items():
        found = None
        for pattern in patterns:
            hits = [i for i, name in enumerate(composite) if re.search(pattern, name)]
            if len(hits) == 1:
                found = hits[0]
                break
            if len(hits) > 1:
                raise RuntimeError(
                    f"Kolonnen '{field}' er tvetydig i {filename}: moenstret "
                    f"'{pattern}' matcher kolonne "
                    + ", ".join(f"{get_column_letter(i + 1)} ({composite[i]!r})" for i in hits)
                )
        if found is None:
            if default_idx >= raw.shape[1]:
                raise RuntimeError(
                    f"Kunne ikke finde kolonnen '{field}' i {filename}, og "
                    f"reservepositionen {get_column_letter(default_idx + 1)} findes ikke. "
                    f"Fundne kolonnenavne: {[c for c in composite if c]}"
                )
            found = default_idx
            warnings.append(
                f"{filename}: kolonnenavnet for '{field}' blev ikke genkendt - "
                f"bruger position {get_column_letter(found + 1)} "
                f"({composite[found] or 'uden navn'!r}) jf. spec."
            )
        elif found != default_idx:
            warnings.append(
                f"{filename}: '{field}' fundet i kolonne "
                f"{get_column_letter(found + 1)}, ikke {get_column_letter(default_idx + 1)} "
                "som specen angiver - filens layout er aendret."
            )
        mapping[field] = found
    return mapping


def detect_layout(path: Path, spec: dict, warnings: list):
    raw = read_table(path)
    if raw.empty:
        raise RuntimeError(f"{path.name} er tom.")
    header_row = find_header_row(raw, spec, path.name)
    data_start = find_data_start(raw, header_row, path.name)
    mapping = map_columns(raw, header_row, data_start, spec, path.name, warnings)
    data = raw.iloc[data_start:].reset_index(drop=True)
    return data, mapping, header_row, data_start


def load_lb(path: Path, warnings: list) -> pd.DataFrame:
    raw, cols, header_row, data_start = detect_layout(path, LB_COLUMNS, warnings)

    # Kritisk kontrol: foerste datalinje SKAL baere et projektnummer. Goer den
    # ikke det, er vi startet for langt nede, og forward-fill vil tavst kaste
    # hele foerste projektblok vaek.
    first_projekt = clean(raw.iat[0, cols["projektnr"]])
    if not first_projekt or "total" in first_projekt.lower():
        raise RuntimeError(
            f"Foerste datalinje i {path.name} (raekke {data_start + 1}) mangler "
            f"projektnummer i kolonne {get_column_letter(cols['projektnr'] + 1)} "
            f"- fandt {first_projekt!r}. Filen ser ud til at starte midt i en "
            "projektblok, og et helt projekt ville blive tabt uden varsel. "
            "Kontroller at hele rapporten er eksporteret."
        )

    df = pd.DataFrame({
        "projektnr": raw.iloc[:, cols["projektnr"]].map(clean),
        "projektnavn": raw.iloc[:, cols["projektnavn"]].map(clean),
        "uk": raw.iloc[:, cols["uk"]].map(clean),
        "kategori": raw.iloc[:, cols["kategori"]].map(clean),
        "lb": raw.iloc[:, cols["lb"]].map(parse_dk_number),
        "actual": raw.iloc[:, cols["actual"]].map(parse_dk_number),
        "difference": raw.iloc[:, cols["difference"]].map(parse_dk_number),
    })

    # Subtotal-etiketter maa ikke forward-fill'es ned over detailraekker
    for col in ("projektnr", "projektnavn", "uk"):
        df[col] = df[col].mask(df[col].str.contains("Total", case=False, na=False), "")
        df[col] = df[col].replace("", pd.NA).ffill().fillna("")

    # Detailraekker = raekker med udfyldt udgiftskategori
    df = df[df["kategori"] != ""].copy()
    df = df[df["projektnr"] != ""].copy()
    return df


def load_stamdata(path: Path, warnings: list) -> pd.DataFrame:
    raw, cols, _, _ = detect_layout(path, SD_COLUMNS, warnings)
    return pd.DataFrame({
        "projektnr": raw.iloc[:, cols["projektnr"]].map(clean),
        "projektnavn_sd": raw.iloc[:, cols["projektnavn_sd"]].map(clean),
        "sektionsnummer": raw.iloc[:, cols["organisation"]].map(extract_sektionsnummer),
        "projektstatus": raw.iloc[:, cols["projektstatus"]].map(clean),
        "startdato": raw.iloc[:, cols["startdato"]].map(parse_us_date),
        "slutdato": raw.iloc[:, cols["slutdato"]].map(parse_us_date),
        "projektcontroller": raw.iloc[:, cols["projektcontroller"]].map(clean),
    }).query("projektnr != ''")


def build_output(lb: pd.DataFrame, stamdata: pd.DataFrame):
    drift = lb[lb["kategori"].isin(DRIFT_KATEGORIER)].copy()

    keys = ["projektnr", "uk", "kategori"] if GROUP_BY_UK else ["projektnr", "kategori"]
    agg = (
        drift.groupby(keys, as_index=False, dropna=False)
        .agg(projektnavn=("projektnavn", "first"),
             lb_ytd=("lb", "sum"),
             actual_ytd=("actual", "sum"),
             diff_kilde=("difference", "sum"))
    )
    if not GROUP_BY_UK:
        agg["uk"] = ""

    sd = stamdata.drop_duplicates(subset="projektnr").set_index("projektnr")
    agg = agg.join(sd, on="projektnr", rsuffix="_sd")

    kat_order = {k: i for i, k in enumerate(DRIFT_KATEGORIER)}
    agg["_kat"] = agg["kategori"].map(kat_order)
    agg = agg.sort_values(
        ["sektionsnummer", "projektnr", "uk", "_kat"], na_position="last"
    ).drop(columns="_kat").reset_index(drop=True)

    lb_projekter = set(lb["projektnr"])
    output_projekter = set(agg["projektnr"])
    fravalgt = []

    for _, row in stamdata.iterrows():
        if row["projektnr"] not in lb_projekter:
            fravalgt.append({
                "Projektnummer": row["projektnr"],
                "Projektnavn": row["projektnavn_sd"],
                "Sektionsnummer": row["sektionsnummer"],
                "Projektcontroller": row["projektcontroller"],
                "Årsag": "Findes i Projektstamdata, men ikke i LB-filen",
            })

    for projektnr in sorted(lb_projekter - output_projekter):
        block = lb[lb["projektnr"] == projektnr]
        kats = sorted(set(block["kategori"]))
        fravalgt.append({
            "Projektnummer": projektnr,
            "Projektnavn": block["projektnavn"].iloc[0],
            "Sektionsnummer": sd["sektionsnummer"].get(projektnr, ""),
            "Projektcontroller": sd["projektcontroller"].get(projektnr, ""),
            "Årsag": "Ingen driftskategorier i LB-filen. Kategorier: " + ", ".join(kats),
        })

    mangler_stamdata = sorted(
        set(agg.loc[agg["sektionsnummer"].isna(), "projektnr"])
        | set(agg.loc[agg["sektionsnummer"] == "", "projektnr"])
    )
    for projektnr in mangler_stamdata:
        fravalgt.append({
            "Projektnummer": projektnr,
            "Projektnavn": agg.loc[agg["projektnr"] == projektnr, "projektnavn"].iloc[0],
            "Sektionsnummer": "",
            "Projektcontroller": "",
            "Årsag": "ADVARSEL: findes i LB-filen, men ikke i Projektstamdata "
                     "- kolonne A, M, N og O er tomme i output",
        })

    return agg, pd.DataFrame(fravalgt)


def style_header(cell):
    cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_workbook(agg, fravalgt, benchmark, month, lb_path, sd_path, out_path):
    wb = Workbook()

    ws_a = wb.active
    ws_a.title = "Forudsætninger"
    ws_a.column_dimensions["A"].width = 34
    ws_a.column_dimensions["B"].width = 62

    rows = [
        ("Controlling af drift i bottom up (LB)", ""),
        ("", ""),
        ("Benchmark (måned/12)", None),          # B3 = inputcelle
        ("Benchmarkmåned", month or ""),
        ("Kørt den", date.today()),
        ("", ""),
        ("LB-fil", lb_path.name),
        ("Projektstamdatafil", sd_path.name),
        ("", ""),
        ("Medtagne udgiftskategorier", "; ".join(DRIFT_KATEGORIER)),
        ("Gruppering", "Projekt + UK + udgiftskategori" if GROUP_BY_UK
                       else "Projekt + udgiftskategori (UK lagt sammen)"),
        ("Antal linjer i output", len(agg)),
        ("", ""),
        ("Bemærk", "Gul celle B3 er eneste inputcelle. Ændres benchmark her, "
                   "genberegnes kolonne J og K på fanen Drift."),
        ("Kolonne H", "Beregnes som F-G (svarer til LB-filens kolonne J)."),
        ("Kolonne I", "G/F, dvs. forbrugsprocent af løbende budget."),
        ("Kolonne K", "G/J-F, dvs. estimeret helårsforbrug minus løbende budget."),
    ]
    for i, (label, value) in enumerate(rows, start=1):
        ws_a.cell(i, 1, label).font = Font(name=FONT_NAME, size=10, bold=(i == 1 or label != ""))
        if value is not None:
            ws_a.cell(i, 2, value).font = Font(name=FONT_NAME, size=10)
    ws_a["A1"].font = Font(name=FONT_NAME, size=13, bold=True)

    b3 = ws_a["B3"]
    b3.value = benchmark
    b3.number_format = "0.00"
    b3.font = Font(name=FONT_NAME, size=10, bold=True, color="FF0000FF")
    b3.fill = ASSUMPTION_FILL
    b3.border = Border(*[Side(style="thin")] * 4)
    ws_a["B5"].number_format = "dd-mm-yyyy"

    ws = wb.create_sheet("Drift")
    for col, header in enumerate(OUTPUT_HEADERS, start=1):
        style_header(ws.cell(1, col, header))
    ws.row_dimensions[1].height = 30

    thin = Side(style="thin", color="FFD9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, row in enumerate(agg.itertuples(index=False), start=2):
        ws.cell(i, 1, row.sektionsnummer if isinstance(row.sektionsnummer, str) else "")
        ws.cell(i, 2, row.projektnr)
        ws.cell(i, 3, row.projektnavn)
        ws.cell(i, 4, row.uk)
        ws.cell(i, 5, row.kategori)
        ws.cell(i, 6, round(row.lb_ytd, 2))
        ws.cell(i, 7, round(row.actual_ytd, 2))
        ws.cell(i, 8, f"=F{i}-G{i}")
        ws.cell(i, 9, f'=IFERROR(G{i}/F{i},"")')
        ws.cell(i, 10, "=Forudsætninger!$B$3")
        ws.cell(i, 11, f'=IFERROR(G{i}/J{i}-F{i},"")')
        ws.cell(i, 12, f'=IF(F{i}=0,"LB = 0","")')
        ws.cell(i, 13, row.projektcontroller if isinstance(row.projektcontroller, str) else "")
        ws.cell(i, 14, row.startdato if row.startdato is not None
                       and not pd.isna(row.startdato) else "")
        ws.cell(i, 15, row.slutdato if row.slutdato is not None
                       and not pd.isna(row.slutdato) else "")

    last = ws.max_row
    if last >= 2:
        for i in range(2, last + 1):
            for col in range(1, 16):
                cell = ws.cell(i, col)
                cell.font = Font(name=FONT_NAME, size=10)
                cell.border = border
            for col in (6, 7, 8, 11):
                ws.cell(i, col).number_format = "#,##0"
            for col in (9, 10):
                ws.cell(i, col).number_format = "0.00"
            for col in (14, 15):
                ws.cell(i, col).number_format = "dd-mm-yyyy"
            ws.cell(i, 2).alignment = Alignment(horizontal="left")
            ws.cell(i, 4).alignment = Alignment(horizontal="center")

        total = last + 1
        ws.cell(total, 5, "I alt").font = Font(name=FONT_NAME, size=10, bold=True)
        for col in (6, 7, 8, 11):
            letter = get_column_letter(col)
            cell = ws.cell(total, col, f"=SUM({letter}2:{letter}{last})")
            cell.font = Font(name=FONT_NAME, size=10, bold=True)
            cell.number_format = "#,##0"
        cell = ws.cell(total, 9, f'=IFERROR(G{total}/F{total},"")')
        cell.font = Font(name=FONT_NAME, size=10, bold=True)
        cell.number_format = "0.00"
        for col in range(1, 16):
            ws.cell(total, col).border = Border(top=Side(style="medium"))

        rng = f"A2:O{last}" if RED_WHOLE_ROW else f"H2:H{last}"
        ws.conditional_formatting.add(
            rng, FormulaRule(formula=[f"$H2<0"], font=Font(color=RED), stopIfTrue=False)
        )

        ws.auto_filter.ref = f"A1:O{last}"

    ws.freeze_panes = "F2"
    for col, width in zip(range(1, 16),
                          [14, 14, 34, 7, 24, 18, 15, 15, 15, 11, 18, 26, 26, 12, 12]):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws_f = wb.create_sheet("Fravalgt")
    cols = ["Projektnummer", "Projektnavn", "Sektionsnummer", "Projektcontroller", "Årsag"]
    for col, header in enumerate(cols, start=1):
        style_header(ws_f.cell(1, col, header))
    for i, row in enumerate(fravalgt.to_dict("records") if len(fravalgt) else [], start=2):
        for col, key in enumerate(cols, start=1):
            cell = ws_f.cell(i, col, row.get(key, ""))
            cell.font = Font(name=FONT_NAME, size=10,
                             color=RED if "ADVARSEL" in str(row.get("Årsag", "")) else "FF000000")
    if not len(fravalgt):
        ws_f.cell(2, 1, "Ingen fravalgte projekter.").font = Font(name=FONT_NAME, size=10, italic=True)
    for col, width in zip(range(1, 6), [16, 38, 16, 28, 74]):
        ws_f.column_dimensions[get_column_letter(col)].width = width
    ws_f.freeze_panes = "A2"

    wb.save(out_path)


def main():
    p = argparse.ArgumentParser(
        description="Controlling af drift i bottom up (LB) - genererer Excel-output.")
    p.add_argument("--lb", required=True, type=Path, help="Sti til LB-filen (csv/xlsx)")
    p.add_argument("--stamdata", required=True, type=Path,
                   help="Sti til Projektstamdatafilen (csv/xlsx)")
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--month", type=int, choices=range(1, 13), metavar="1-12",
                   help="Ultimo hvilken kalendermaaned data daekker. Benchmark = maaned/12")
    g.add_argument("--benchmark", type=float,
                   help="Benchmark angivet direkte, fx 0.42")
    p.add_argument("--output", type=Path, default=Path("Drift_output.xlsx"))
    args = p.parse_args()

    for path in (args.lb, args.stamdata):
        if not path.exists():
            sys.exit(f"FEJL: filen findes ikke: {path}")

    benchmark = args.benchmark if args.benchmark is not None else args.month / 12

    warnings: list = []
    lb = load_lb(args.lb, warnings)
    stamdata = load_stamdata(args.stamdata, warnings)
    agg, fravalgt = build_output(lb, stamdata)
    write_workbook(agg, fravalgt, benchmark, args.month,
                   args.lb, args.stamdata, args.output)

    label = f"ultimo {MAANEDSNAVNE[args.month]}" if args.month else "manuelt angivet"
    print(f"LB-detailraekker indlaest : {len(lb)}")
    print(f"Heraf driftskategorier    : {len(lb[lb['kategori'].isin(DRIFT_KATEGORIER)])}")
    print(f"Outputlinjer              : {len(agg)}")
    print(f"Fravalgte/advarsler       : {len(fravalgt)}")
    print(f"Benchmark                 : {benchmark:.4f} ({label})")
    print(f"Skrevet                   : {args.output}")
    if warnings:
        print("\nADVARSLER:")
        for w in warnings:
            print(f"  ! {w}")


if __name__ == "__main__":
    main()
