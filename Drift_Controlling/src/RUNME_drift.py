#!/usr/bin/env python3
"""Controlling af drift i bottom up (LB). Baggrund og datakvirks: se CLAUDE.md."""

import argparse
import csv
import io
import re
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

GROUP_BY_UK = True
RED_WHOLE_ROW = True

# Udgiftskategorier der regnes som drift. Alt andet frasorteres.
DRIFT_KATEGORIER = [
    "Intern produktion mv., køb", "Lejeomkostninger o.l.",
    "Rejser og repræsentation", "Øvrige omkostninger",
]

# Fanen Drift, kolonne A-O
OUTPUT_HEADERS = [
    "Sektionsnummer", "Projektnummer", "Projektnavn", "UK", "Udgiftskategori",
    "Løbende budget YTD", "Actuals YTD", "Difference YTD", "Difference YTD %",
    "Benchmark", "Beregnet justering", "Kommentar YTD", "Projektcontroller",
    "Startdato", "Slutdato",
]
BREDDER = [14, 14, 34, 7, 24, 18, 15, 15, 15, 11, 18, 26, 26, 12, 12]
FORMATER = {6: "#,##0", 7: "#,##0", 8: "#,##0", 11: "#,##0",
            9: "0.00", 10: "0.00", 14: "dd-mm-yyyy", 15: "dd-mm-yyyy"}
FRAVALGT_COLS = ["Projektnummer", "Projektnavn", "Sektionsnummer",
                 "Projektcontroller", "Årsag"]

MAANEDER = ("januar februar marts april maj juni juli august september "
            "oktober november december").split()

FONT_NAME = "Arial"
RED = "FFC00000"
F10 = Font(name=FONT_NAME, size=10)
F10B = Font(name=FONT_NAME, size=10, bold=True)
F10SORT = Font(name=FONT_NAME, size=10, color="FF000000")
F10ROED = Font(name=FONT_NAME, size=10, color=RED)
HEADER_FILL = PatternFill("solid", fgColor="FF1F3864")
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFFFF00")

NUMERIC_RE = re.compile(r"^-?\(?[\d.\s ]*\d(?:,\d+)?\)?-?$")

# felt -> (forventet position jf. spec, moenstre der matcher kolonnenavnet)
LB_COLUMNS = {
    "projektnr":   (0, [r"project\s*no", r"projektnr"]),
    "projektnavn": (1, [r"projektnavn", r"project\s*name"]),
    "uk":          (2, [r"^uk$", r"underkonto"]),
    "kategori":    (6, [r"udgiftskategori", r"expenditure\s*category"]),
    "lb":          (7, [r"forecast\s*raw", r"l[øo]bende\s*budget"]),
    "actual":      (8, [r"actual\s*raw"]),
    "difference":  (9, [r"difference\s*raw"]),
}
SD_COLUMNS = {
    "projektnr":         (0, [r"projektnr"]),
    "projektnavn_sd":    (1, [r"projektnavn"]),
    "organisation":      (3, [r"project\s*organi", r"projektorgani"]),
    "projektstatus":     (4, [r"projektstatus"]),
    "startdato":         (5, [r"projektstartdato", r"startdato"]),
    "slutdato":          (6, [r"projektslutdato", r"slutdato"]),
    "projektcontroller": (8, [r"project\s*controller", r"projektcontroller"]),
}


# --------------------------------------------------------------------------- #
# Indlaesning og parsing
# --------------------------------------------------------------------------- #

def read_table(path: Path) -> pd.DataFrame:
    # xlsx: dtype=object, ALDRIG dtype=str. dtype=str goer floaten 984935.88 til
    # teksten "984935.88", som den danske talparser laeser som 98.493.588.131.313,
    # og datoen til "2021-12-01 00:00:00", som intet datoformat matcher.
    if path.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
        raw = pd.read_excel(path, header=None, dtype=object)
        return raw[raw.map(clean).ne("").any(axis=1)].reset_index(drop=True)

    text = None
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = path.read_text(encoding=encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if text is None:
        raise RuntimeError(f"Kunne ikke afkode {path.name}.")

    try:
        delimiter = csv.Sniffer().sniff(
            "\n".join(text.splitlines()[:20]), delimiters=";,\t").delimiter
    except csv.Error:
        delimiter = ";"

    rows = [r for r in csv.reader(io.StringIO(text), delimiter=delimiter)
            if any(c.strip() for c in r)]
    if not rows:
        raise RuntimeError(f"{path.name} indeholder ingen raekker.")
    width = max(len(r) for r in rows)
    return pd.DataFrame([r + [""] * (width - len(r)) for r in rows], dtype=str)


def detect_number_style(values):
    """Dansk ('.' tusind, ',' decimal) eller engelsk ('.' decimal)?

    Returnerer (style, dk_hits, us_hits). Rigtige tal fra en xlsx springes
    over - kun tekst kan vaere tvetydig, og kun tekst stemmer derfor.
    """
    dk = us = 0
    for value in values:
        if not isinstance(value, str):
            continue
        text = value.strip().strip("()").replace("\xa0", "").replace(" ", "")
        if not text or not any(c.isdigit() for c in text):
            continue
        if "," in text or text.count(".") >= 2:
            dk += 1                 # ',' kan kun vaere decimal; '1.234.567' = tusindtal
        elif text.count(".") == 1 and len(text.rsplit(".", 1)[1]) != 3:
            us += 1                 # '984935.88' = decimal
    # '599.665' alene er reelt tvetydigt - specen siger dansk, saa dansk vinder.
    return ("us" if us and not dk else "dk"), dk, us


def parse_number(value, style: str = "dk") -> float:
    """'2.782.094,86' -> 2782094.86 (dansk), '984935.88' -> 984935.88 (engelsk).
    Tal fra en xlsx gaar igennem uroert. Tomme felter -> 0.0"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text in ("-", "nan"):
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace("\xa0", "").replace(" ", "")
    text = text.replace(",", "") if style == "us" else \
        text.replace(".", "").replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return 0.0
    return -number if negative else number


def parse_us_date(value):
    """Projektstamdata leverer datoer som MM-DD-YYYY (fx 10-30-2023)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (pd.Timestamp, date)):
        return pd.Timestamp(value).date()
    # Som tekst haenger klokkeslaettet tit med ('2021-12-01 00:00:00')
    text = re.sub(r"[\sT]+\d{1,2}:\d{2}(:\d{2})?(\.\d+)?$", "", str(value).strip()).strip()
    for fmt in ("%m-%d-%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return pd.to_datetime(text, format=fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def extract_sektionsnummer(value) -> str:
    """'Afdeling for ... , (7810) Administration' -> '7810'"""
    text = clean(value)
    match = re.search(r"\((\d{4})\)", text) or re.search(r"(?<!\d)(\d{4})(?!\d)", text)
    return match.group(1) if match else ""


def clean(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    # Et projektnummer eller en UK, der ligger som tal i en xlsx, maa ikke
    # blive til '100196.0' - saa fejler opslaget mod Projektstamdata.
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def norm(value) -> str:
    return clean(value).lower()


def looks_numeric(value) -> bool:
    text = norm(value)
    return bool(text) and bool(NUMERIC_RE.match(text)) and any(c.isdigit() for c in text)


def cell_value(value):
    """NaN/None -> '' saa openpyxl ikke skriver 'nan' i regnearket."""
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return ""
    return value


# --------------------------------------------------------------------------- #
# Layoutgenkendelse - antallet af headerraekker er ikke fast, se CLAUDE.md
# --------------------------------------------------------------------------- #

def detect_layout(path: Path, spec: dict, warnings: list):
    raw = read_table(path)
    if raw.empty:
        raise RuntimeError(f"{path.name} er tom.")
    navn = path.name

    # Headerraekken = den raekke hvor flest kolonnenavne genkendes
    header_row, best_hits = None, 0
    for i in range(min(25, len(raw))):
        cells = [norm(v) for v in raw.iloc[i]]
        hits = sum(any(re.search(p, c) for c in cells for p in patterns)
                   for _, patterns in spec.values())
        if hits > best_hits:
            header_row, best_hits = i, hits
    if header_row is None or best_hits < 3:
        raise RuntimeError(
            f"Kunne ikke finde headerraekken i {navn}. Bedste kandidat genkendte "
            f"{best_hits} af {len(spec)} kolonnenavne. Er filen semikolonsepareret "
            "og eksporteret fra den rigtige rapport?")

    # Datastart = foerste raekke efter headeren med et rigtigt tal. Tekstlige
    # fortsaettelsesraekker springes over, uanset hvor mange der er.
    data_start = next((i for i in range(header_row + 1, len(raw))
                       if any(looks_numeric(v) for v in raw.iloc[i])), None)
    if data_start is None:
        raise RuntimeError(
            f"Fandt ingen datalinjer i {navn} efter headerraekke {header_row + 1}.")

    # Sammensatte kolonnenavne (header -> datastart), saa moenstre som
    # 'forecast\s*raw' ogsaa matcher en header fordelt over to raekker.
    composite = [norm(" ".join(filter(None, (clean(raw.iat[r, col])
                                             for r in range(header_row, data_start)))))
                 for col in range(raw.shape[1])]

    mapping = {}
    for field, (default_idx, patterns) in spec.items():
        found = None
        for pattern in patterns:
            hits = [i for i, name in enumerate(composite) if re.search(pattern, name)]
            if len(hits) == 1:
                found = hits[0]
                break
            if len(hits) > 1:
                raise RuntimeError(
                    f"Kolonnen '{field}' er tvetydig i {navn}: moenstret "
                    f"'{pattern}' matcher kolonne " + ", ".join(
                        f"{get_column_letter(i + 1)} ({composite[i]!r})" for i in hits))
        if found is None:
            if default_idx >= raw.shape[1]:
                raise RuntimeError(
                    f"Kunne ikke finde kolonnen '{field}' i {navn}, og "
                    f"reservepositionen {get_column_letter(default_idx + 1)} findes "
                    f"ikke. Fundne kolonnenavne: {[c for c in composite if c]}")
            found = default_idx
            warnings.append(
                f"{navn}: kolonnenavnet for '{field}' blev ikke genkendt - bruger "
                f"position {get_column_letter(found + 1)} "
                f"({composite[found] or 'uden navn'!r}) jf. spec.")
        elif found != default_idx:
            warnings.append(
                f"{navn}: '{field}' fundet i kolonne {get_column_letter(found + 1)}, "
                f"ikke {get_column_letter(default_idx + 1)} som specen angiver - "
                "filens layout er aendret.")
        mapping[field] = found

    return raw.iloc[data_start:].reset_index(drop=True), mapping, data_start


def load_lb(path: Path, warnings: list) -> pd.DataFrame:
    raw, cols, data_start = detect_layout(path, LB_COLUMNS, warnings)

    # Kritisk kontrol: foerste datalinje SKAL baere et projektnummer. Goer den
    # ikke det, er vi startet for langt nede, og forward-fill vil tavst kaste
    # hele foerste projektblok vaek.
    first_projekt = clean(raw.iat[0, cols["projektnr"]])
    if not first_projekt or "total" in first_projekt.lower():
        raise RuntimeError(
            f"Foerste datalinje i {path.name} (raekke {data_start + 1}) mangler "
            f"projektnummer i kolonne {get_column_letter(cols['projektnr'] + 1)} "
            f"- fandt {first_projekt!r}. Filen ser ud til at starte midt i en "
            "projektblok, og et helt projekt ville blive tabt uden varsel. "
            "Kontroller at hele rapporten er eksporteret.")

    # Talformatet afgoeres paa filniveau, ikke pr. kolonne: en enkelt kolonne
    # kan bestaa af rene heltal og dermed vaere formatloes.
    style, dk_hits, us_hits = detect_number_style(
        pd.concat([raw.iloc[:, cols[f]] for f in ("lb", "actual", "difference")]))
    if style == "us":
        warnings.append(
            f"{path.name}: talkolonnerne bruger '.' som decimalseparator (engelsk "
            f"format, {us_hits} vaerdier) og laeses saadan. Specen beskriver dansk "
            "format - kontroller stoerrelsesordenen i output.")
    elif dk_hits and us_hits:
        warnings.append(
            f"{path.name}: blandet talformat - {dk_hits} vaerdier ser danske ud og "
            f"{us_hits} engelske. Laest som dansk. Kontroller output.")

    df = pd.DataFrame({
        felt: raw.iloc[:, cols[felt]].map(clean)
        for felt in ("projektnr", "projektnavn", "uk", "kategori")
    } | {
        felt: raw.iloc[:, cols[felt]].map(lambda v: parse_number(v, style))
        for felt in ("lb", "actual", "difference")
    })

    # Subtotal-etiketter maa ikke forward-fill'es ned over detailraekker
    for col in ("projektnr", "projektnavn", "uk"):
        df[col] = df[col].mask(df[col].str.contains("Total", case=False, na=False), "")
        df[col] = df[col].replace("", pd.NA).ffill().fillna("")

    # Detailraekker = raekker med udfyldt udgiftskategori
    return df[(df["kategori"] != "") & (df["projektnr"] != "")].copy()


def load_stamdata(path: Path, warnings: list) -> pd.DataFrame:
    raw, cols, _ = detect_layout(path, SD_COLUMNS, warnings)
    col = lambda felt: raw.iloc[:, cols[felt]]  # noqa: E731
    return pd.DataFrame({
        "projektnr": col("projektnr").map(clean),
        "projektnavn_sd": col("projektnavn_sd").map(clean),
        "sektionsnummer": col("organisation").map(extract_sektionsnummer),
        "projektstatus": col("projektstatus").map(clean),
        "startdato": col("startdato").map(parse_us_date),
        "slutdato": col("slutdato").map(parse_us_date),
        "projektcontroller": col("projektcontroller").map(clean),
    }).query("projektnr != ''")


# --------------------------------------------------------------------------- #
# Aggregering
# --------------------------------------------------------------------------- #

def build_output(lb: pd.DataFrame, stamdata: pd.DataFrame):
    drift = lb[lb["kategori"].isin(DRIFT_KATEGORIER)].copy()

    keys = ["projektnr", "uk", "kategori"] if GROUP_BY_UK else ["projektnr", "kategori"]
    agg = drift.groupby(keys, as_index=False, dropna=False).agg(
        projektnavn=("projektnavn", "first"), lb_ytd=("lb", "sum"),
        actual_ytd=("actual", "sum"), diff_kilde=("difference", "sum"))
    if not GROUP_BY_UK:
        agg["uk"] = ""

    sd = stamdata.drop_duplicates(subset="projektnr").set_index("projektnr")
    agg = agg.join(sd, on="projektnr", rsuffix="_sd")

    agg["_kat"] = agg["kategori"].map({k: i for i, k in enumerate(DRIFT_KATEGORIER)})
    agg = agg.sort_values(["sektionsnummer", "projektnr", "uk", "_kat"],
                          na_position="last").drop(columns="_kat").reset_index(drop=True)

    fravalgt = []

    def tilfoej(nr, navn, sektion, controller, aarsag):
        fravalgt.append(dict(zip(FRAVALGT_COLS, (nr, navn, sektion, controller, aarsag))))

    lb_projekter = set(lb["projektnr"])
    for _, row in stamdata.iterrows():
        if row["projektnr"] not in lb_projekter:
            tilfoej(row["projektnr"], row["projektnavn_sd"], row["sektionsnummer"],
                    row["projektcontroller"],
                    "Findes i Projektstamdata, men ikke i LB-filen")

    for projektnr in sorted(lb_projekter - set(agg["projektnr"])):
        block = lb[lb["projektnr"] == projektnr]
        tilfoej(projektnr, block["projektnavn"].iloc[0],
                sd["sektionsnummer"].get(projektnr, ""),
                sd["projektcontroller"].get(projektnr, ""),
                "Ingen driftskategorier i LB-filen. Kategorier: "
                + ", ".join(sorted(set(block["kategori"]))))

    for projektnr in sorted(set(agg.loc[agg["sektionsnummer"].isna(), "projektnr"])
                            | set(agg.loc[agg["sektionsnummer"] == "", "projektnr"])):
        tilfoej(projektnr, agg.loc[agg["projektnr"] == projektnr, "projektnavn"].iloc[0],
                "", "", "ADVARSEL: findes i LB-filen, men ikke i Projektstamdata "
                        "- kolonne A, M, N og O er tomme i output")

    return agg, pd.DataFrame(fravalgt)


# --------------------------------------------------------------------------- #
# Workbook
# --------------------------------------------------------------------------- #

def style_header(cell):
    cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _forudsaetninger(wb, agg, benchmark, month, lb_path, sd_path):
    ws = wb.active
    ws.title = "Forudsætninger"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 62

    for i, (label, value) in enumerate([
        ("Controlling af drift i bottom up (LB)", ""),
        ("", ""),
        ("Benchmark (måned/12)", None),          # B3 = inputcelle
        ("Benchmarkmåned", month or ""),
        ("Kørt den", date.today()),
        ("", ""),
        ("LB-fil", lb_path.name),
        ("Projektstamdatafil", sd_path.name),
        ("", ""),
        ("Medtagne udgiftskategorier", "; ".join(DRIFT_KATEGORIER)),
        ("Gruppering", "Projekt + UK + udgiftskategori" if GROUP_BY_UK
                       else "Projekt + udgiftskategori (UK lagt sammen)"),
        ("Antal linjer i output", len(agg)),
        ("", ""),
        ("Bemærk", "Gul celle B3 er eneste inputcelle. Ændres benchmark her, "
                   "genberegnes kolonne J og K på fanen Drift."),
        ("Kolonne H", "Beregnes som F-G (svarer til LB-filens kolonne J)."),
        ("Kolonne I", "G/F, dvs. forbrugsprocent af løbende budget."),
        ("Kolonne K", "G/J-F, dvs. estimeret helårsforbrug minus løbende budget."),
    ], start=1):
        ws.cell(i, 1, label).font = F10B if (i == 1 or label != "") else F10
        if value is not None:
            ws.cell(i, 2, value).font = F10
    ws["A1"].font = Font(name=FONT_NAME, size=13, bold=True)

    b3 = ws["B3"]
    b3.value = benchmark
    b3.number_format = "0.00"
    b3.font = Font(name=FONT_NAME, size=10, bold=True, color="FF0000FF")
    b3.fill = ASSUMPTION_FILL
    b3.border = Border(*[Side(style="thin")] * 4)
    ws["B5"].number_format = "dd-mm-yyyy"


def _drift(wb, agg):
    ws = wb.create_sheet("Drift")
    for col, header in enumerate(OUTPUT_HEADERS, start=1):
        style_header(ws.cell(1, col, header))
    ws.row_dimensions[1].height = 30

    for i, row in enumerate(agg.itertuples(index=False), start=2):
        for col, value in enumerate([
            cell_value(row.sektionsnummer), row.projektnr, row.projektnavn,
            row.uk, row.kategori, round(row.lb_ytd, 2), round(row.actual_ytd, 2),
            f"=F{i}-G{i}", f'=IFERROR(G{i}/F{i},"")', "=Forudsætninger!$B$3",
            f'=IFERROR(G{i}/J{i}-F{i},"")', f'=IF(F{i}=0,"LB = 0","")',
            cell_value(row.projektcontroller), cell_value(row.startdato),
            cell_value(row.slutdato),
        ], start=1):
            ws.cell(i, col, value)

    last = ws.max_row
    if last >= 2:
        thin = Side(style="thin", color="FFD9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        venstre, midt = Alignment(horizontal="left"), Alignment(horizontal="center")
        for i in range(2, last + 1):
            for col in range(1, 16):
                cell = ws.cell(i, col)
                cell.font, cell.border = F10, border
                if col in FORMATER:
                    cell.number_format = FORMATER[col]
            ws.cell(i, 2).alignment = venstre
            ws.cell(i, 4).alignment = midt

        total = last + 1
        ws.cell(total, 5, "I alt").font = F10B
        for col in (6, 7, 8, 11, 9):
            letter = get_column_letter(col)
            cell = ws.cell(total, col, f'=IFERROR(G{total}/F{total},"")' if col == 9
                           else f"=SUM({letter}2:{letter}{last})")
            cell.font, cell.number_format = F10B, FORMATER[col]
        for col in range(1, 16):
            ws.cell(total, col).border = Border(top=Side(style="medium"))

        ws.conditional_formatting.add(
            f"A2:O{last}" if RED_WHOLE_ROW else f"H2:H{last}",
            FormulaRule(formula=["$H2<0"], font=Font(color=RED), stopIfTrue=False))
        ws.auto_filter.ref = f"A1:O{last}"

    ws.freeze_panes = "F2"
    for col, width in enumerate(BREDDER, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _fravalgt(wb, fravalgt):
    ws = wb.create_sheet("Fravalgt")
    for col, header in enumerate(FRAVALGT_COLS, start=1):
        style_header(ws.cell(1, col, header))
    for i, row in enumerate(fravalgt.to_dict("records") if len(fravalgt) else [], start=2):
        roed = "ADVARSEL" in str(row.get("Årsag", ""))
        for col, key in enumerate(FRAVALGT_COLS, start=1):
            ws.cell(i, col, row.get(key, "")).font = F10ROED if roed else F10SORT
    if not len(fravalgt):
        ws.cell(2, 1, "Ingen fravalgte projekter.").font = Font(
            name=FONT_NAME, size=10, italic=True)
    for col, width in enumerate([16, 38, 16, 28, 74], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def write_workbook(agg, fravalgt, benchmark, month, lb_path, sd_path, out_path):
    wb = Workbook()
    _forudsaetninger(wb, agg, benchmark, month, lb_path, sd_path)
    _drift(wb, agg)
    _fravalgt(wb, fravalgt)
    wb.save(out_path)


# --------------------------------------------------------------------------- #
# Koersel
# --------------------------------------------------------------------------- #

def print_summary(lb, agg, fravalgt, benchmark, month, out_path, warnings):
    label = f"ultimo {MAANEDER[month - 1]}" if month else "manuelt angivet"
    for felt, vaerdi in (
        ("LB-detailraekker indlaest", len(lb)),
        ("Heraf driftskategorier", len(lb[lb["kategori"].isin(DRIFT_KATEGORIER)])),
        ("Outputlinjer", len(agg)),
        ("Fravalgte/advarsler", len(fravalgt)),
        ("Benchmark", f"{benchmark:.4f} ({label})"),
        ("Skrevet", out_path),
    ):
        print(f"{felt:<26}: {vaerdi}")
    if warnings:
        print("\nADVARSLER:")
        for w in warnings:
            print(f"  ! {w}")


def run(lb_path: Path, sd_path: Path, benchmark: float, month, out_path: Path):
    """Indlaes, aggreger, skriv workbook og udskriv opsummering."""
    warnings: list = []
    lb = load_lb(lb_path, warnings)
    stamdata = load_stamdata(sd_path, warnings)
    agg, fravalgt = build_output(lb, stamdata)
    write_workbook(agg, fravalgt, benchmark, month, lb_path, sd_path, out_path)
    print_summary(lb, agg, fravalgt, benchmark, month, out_path, warnings)
    return agg, fravalgt, warnings


def main():
    p = argparse.ArgumentParser(
        description="Controlling af drift i bottom up (LB) - genererer Excel-output.")
    p.add_argument("--lb", required=True, type=Path, help="Sti til LB-filen (csv/xlsx)")
    p.add_argument("--stamdata", required=True, type=Path,
                   help="Path til Projektstamdatafilen (csv/xlsx)")
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("--month", type=int, choices=range(1, 13), metavar="1-12",
                   help="Ultimo hvilken kalendermaaned data daekker. Benchmark = maaned/12")
    g.add_argument("--benchmark", type=float, help="Benchmark angivet direkte, fx 0.42")
    p.add_argument("--output", type=Path, default=Path("Drift_output.xlsx"))
    args = p.parse_args()

    for path in (args.lb, args.stamdata):
        if not path.exists():
            sys.exit(f"FEJL: filen findes ikke: {path}")

    benchmark = args.benchmark if args.benchmark is not None else args.month / 12
    run(args.lb, args.stamdata, benchmark, args.month, args.output)


if __name__ == "__main__":
    main()
